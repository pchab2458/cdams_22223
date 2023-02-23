from .forms import PaymentForm, BookRoomForm
from .forms import Elec_cpu_change, Water_cpu_change
from .models import Extra, Room, Room_type, MaintenanceService
from django.utils.dateparse import parse_datetime
from django.utils.timezone import is_aware, make_aware
from django.shortcuts import get_object_or_404
from .forms import BillForm
from my_app.models import Billing
from django.views.generic import TemplateView
from django.contrib.auth.decorators import login_required
from .forms import TenantCreateForm, TenantProfileCreateForm
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
from django.views import generic
from django.shortcuts import render, redirect, reverse
from users.forms import CustomUserCreationForm
from my_app.models import TenantProfile
from django.contrib.auth import get_user_model
import random
import calendar
from datetime import datetime, date, timedelta
import decimal
import GV

from openpyxl import Workbook, load_workbook

import os

CUser = get_user_model()


class CholladaHomePage(TemplateView):
    template_name = 'my_app/Chollada_Apartment.html'  # default template if not defined in the url


@login_required
def gateway(request):
    return render(request, 'my_app/admin_page.html')


@login_required
def admin_page(request):
    # return render(request, 'my_app/admin_page.html')
    return render(request, 'my_app/admin_page.html')


@login_required
def create_contract(request):
    # ---------- create vacant room set --------------
    cur_tpfs = TenantProfile.objects.all()  # QS
    all_rms = Room.objects.all().order_by('room_no')  # QS

    oc_rm_set = []  # set of occupied room objects
    vac_rm_set = []  # set of vacant room objects

    for tp in cur_tpfs:
        oc_rm_set.append(tp.room_no)
    for r in all_rms:
        if r not in oc_rm_set:
            vac_rm_set.append(r)
    # -----------------------------------------------
    if request.method == 'POST':
        tenant_form = TenantCreateForm(data=request.POST)
        tenant_profile_form = TenantProfileCreateForm(data=request.POST, files=request.FILES)

        # new_tenant = None
        if tenant_form.is_valid() and tenant_profile_form.is_valid():

            # Create a new tenant object but avoid saving it yet
            new_tenant = tenant_form.save(commit=False)

            # Set the chosen password
            # new_tenant.set_password(tenant_form.cleaned_data['password'])
            new_tenant.set_password(tenant_form.clean_password2())

            # Save the new_tenant object
            new_tenant.save()

            # Create a new tenantprofile object but avoid saving it yet
            tenant_profile = tenant_profile_form.save(commit=False)  # save_m2m() added to tenant_profile_form

            # Set the chosen tenant field
            tenant_profile.tenant = new_tenant

            # Create Room Obj ------------------------------------------------
            rm_obj = get_object_or_404(Room, room_no=request.POST['room_no'])
            # ----------------------------------------------------------------

            tenant_profile.room_no = rm_obj  # assigned obj to obj

            # provide initial value to certain fields before saving to DB
            tenant_profile.elec_unit = 0
            tenant_profile.water_unit = 0
            tenant_profile.misc_cost = 0

            tenant_profile.bill_ref = ""  # BLANK ... ADDED
            tenant_profile.end_date = tenant_profile.start_date.__add__(timedelta(days=1 * 365))

            # -----------------------------------------------------------

            # Save the tenantprofile object
            tenant_profile.save()

            # Save the ManyToMany
            tenant_profile_form.save_m2m()

            messages.success(request, 'ข้อมูลของผู้เช่า ได้ถูกเอาเข้าระบบแล้ว')
            return render(request, 'my_app/admin_page.html', {'section': 'new_contract'})
        else:
            messages.error(request, 'มีข้อผิดพลาดเกิดขึ้น !!!')
            # ------------------------------------------------------------------------------
            return render(request, 'my_app/admin_page.html', {'section': 'new_contract'})
            # ------------------------------------------------------------------------------
    else:
        tenant_form = TenantCreateForm()
        tenant_profile_form = TenantProfileCreateForm()

        return render(request, 'my_app/create_contract.html',
                      {'section': 'new_contract',
                       'tenant_form': tenant_form,
                       'tenant_profile_form': tenant_profile_form,

                       # -----------------------------
                       'vac_rm_set': vac_rm_set,
                       # -----------------------------

                       }
                      )


# START OF CREATE BILLS ==============================================================================

@login_required
def billing(request):
    # --- Check fo OPEN Bills --------------------------------------------------------
    open_bill = Billing.objects.filter(status='open')  # QuerySet

    if open_bill:
        messages.error(request, 'มี {} บิล ยังค้างอยู่ ปิดให้เรียบร้อยก่อน'.format(len(open_bill)))
        return render(request, 'my_app/admin_page.html', {'section': 'billing'})
    else:

        # ------------------------------------------------------------------------------
        all_bills = Billing.objects.all().order_by('-bill_date')  # QS
        if len(all_bills) != 0:
            tpfs = TenantProfile.objects.all()
            latest_bill = all_bills.first()  # bill obj
            latest_billdate = latest_bill.bill_date

            for i in tpfs:
                if i.bill_ref == "" and i.start_date < latest_billdate:
                    messages.error(request, 'ห้อง {} ไม่มี bill_ref !!! ... ตรวจสอบให้ถูกต้องก่อนดำเนินการต่อ ....'.format(i.room_no.room_no))
                    return render(request, 'my_app/admin_page.html', {'section': 'billing'})

        # ------------------------------------------------------------------------------

        cur_date = datetime.now().date()

        # QuerySet: All tenantprofiles in DB created before current date
        tenant_pf = TenantProfile.objects.filter(start_date__lt=cur_date).order_by("room_no")  # QS

        total_tn = len(tenant_pf)

        tpf_billForm_list = []

        for i in tenant_pf:
            rmn = i.room_no.room_no

            prefix = 'RM' + rmn  # RMA101 etc.

            tpf_billForm_list.append((i, BillForm(prefix=prefix)))  # [(tpf,bf),(tpf,bf), ....]

        no_of_bill = 0

        if request.method == 'POST':
            for tpf in tenant_pf:

                rmn = tpf.room_no.room_no
                prefix = "RM" + rmn

                bill_form = BillForm(data=request.POST, instance=tpf, prefix=prefix)

                if bill_form.is_valid():

                    bill_form.save(commit=True)  # CREATE TenanatProfile instance & saved/update ?

                    # -----------------------
                    create_bill(request, rmn)
                    no_of_bill += 1
                    # -----------------------
                else:

                    messages.error(request, 'มีข้อผิดพลาด ไม่สามารถเปิดบิลห้อง {} ได้'.format(rmn))

            messages.success(request, 'เปิด {} บิล จากทั้งหมด {} บิล เรียบร้อยแล้ว'.format(no_of_bill, total_tn))

            return render(request, 'my_app/admin_page.html', {'section': 'billing'})


        else:

            return render(request, 'my_app/billing.html',
                          {
                              'tenant_pf': tenant_pf,
                              'section': 'billing',
                              'cur_date': cur_date,
                              'total_tn': total_tn,
                              'tpf_billForm_list': tpf_billForm_list
                          })


@login_required
def create_bill(req, room_no):
    pf = get_object_or_404(TenantProfile, room_no__room_no=room_no)
    tname = pf.tenant.first_name + ' ' + pf.tenant.last_name

    rno = pf.room_no.room_no
    adj = pf.adjust

    exd = {}
    exd.setdefault('Electricity CPU', 0)
    exd.setdefault('Water CPU', 0)
    exd.setdefault('Garbage', 0)
    exd.setdefault('Parking', 0)
    exd.setdefault('Wifi', 0)

    exd.setdefault('Bed&Mattress', 0)

    exd.setdefault('Dressing Table', 0)
    exd.setdefault('Clothing Cupboard', 0)
    exd.setdefault('TV Table', 0)
    exd.setdefault('Fridge', 0)
    exd.setdefault('Air-Conditioner', 0)

    for e in pf.extra.all():
        exd.update({e.description: e.cpu})

    room_cost = pf.room_no.room_type.rate
    room_acc_cost = exd['Bed&Mattress'] + exd['Dressing Table'] \
                    + exd['Clothing Cupboard'] + exd['TV Table'] + exd['Fridge'] \
                    + exd['Air-Conditioner']

    elec_cost = exd['Electricity CPU'] * pf.elec_unit
    water_cost = exd['Water CPU'] * pf.water_unit

    com_ser_cost = pf.elec_unit * GV.COMMOM_SERVICE_CPU

    oth_ser_cost = exd['Garbage'] + exd['Parking'] + exd['Wifi']

    ovd_amt = pf.cum_ovd

    # -----------------------
    late_f = pf.late_fee
    maint_c = pf.maint_cost

    # RESET pf.late_fee & pf.maint_cost TO O TO BE READY FOR NEXT CYCLE
    pf.late_fee = 0
    pf.maint_cost = 0
    # -----------------------

    total = room_cost + room_acc_cost + elec_cost + water_cost + com_ser_cost + oth_ser_cost + ovd_amt + adj + late_f + maint_c

    # CREATE PRELIMINARY BILL OBJECT **************
    multiplied_factor = 1
    if pf.bill_ref == "":
        if pf.start_date.month == date.today().month:
            multiplied_factor = (calendar.monthrange(date.today().year, date.today().month)[1] - pf.start_date.day) / 30
        else:
            multiplied_factor = (calendar.monthrange(date.today().year, date.today().month)[1] + (calendar.monthrange(pf.start_date.year, pf.start_date.month)[1] - pf.start_date.day)) / 30

    bill_ref = get_ref_string()
    pf.bill_ref = bill_ref  # SET WITH NEW Bill_Ref

    new_bill = Billing(bill_ref=bill_ref,  # CREATE Billing Object
                       bill_date=datetime.now().date(),  # SUPPLY BILL DATE

                       tenant_name=tname,
                       room_no=rno,
                       room_cost=room_cost,
                       room_acc_cost=room_acc_cost,
                       electricity_cost=elec_cost,
                       water_cost=water_cost,
                       common_ser_cost=com_ser_cost,
                       other_ser_cost=oth_ser_cost,
                       overdue_amount=ovd_amt,

                       # -----------------------
                       late_fee=late_f,
                       maint_cost=maint_c,
                       # -----------------------

                       adjust=adj,
                       bill_total=total,

                       )

    # SAVE TENANTPROFILE OBJECT TO DB
    pf.save()  # The tpf CONTAINS bill_ref, elec. units, water units etc.

    # ADJUST PRELIMINARY BILL OBJECT
    adjust_bill(req, pf, new_bill, decimal.Decimal(multiplied_factor))  # ADDED


@login_required
def adjust_bill(req, pf, new_bill, mult_factor):
    tn_bill = new_bill

    bref = tn_bill.bill_ref

    bdate = tn_bill.bill_date

    # bupd # TO BE FILLED WHEN SAVED
    # bstat # TO BE FILLED WHEN SAVED
    tname = tn_bill.tenant_name
    rno = tn_bill.room_no
    room_cost = tn_bill.room_cost
    room_acc_cost = tn_bill.room_acc_cost
    elec_cost = tn_bill.electricity_cost
    water_cost = tn_bill.water_cost
    com_ser_cost = tn_bill.common_ser_cost
    oth_ser_cost = tn_bill.other_ser_cost
    ovd_amt = tn_bill.overdue_amount
    adj = tn_bill.adjust
    # total = tn_bill.bill_total # TO BE ADJUSTED IF REQUIRED
    # pay_date # TO BE FILLED AT PAYMENT
    # pay_amt #TO BE FILL AT PAYMENT
    # bf #TO BE FILLED AT PAYMENT

    late_f = tn_bill.late_fee
    maint_c = tn_bill.maint_cost

    room_cost = room_cost * mult_factor
    room_acc_cost = room_acc_cost * mult_factor
    com_ser_cost = com_ser_cost * mult_factor
    oth_ser_cost = oth_ser_cost * mult_factor
    adj = adj * mult_factor

    total = (room_cost + room_acc_cost + adj) + elec_cost + water_cost + (
            com_ser_cost + oth_ser_cost) + ovd_amt + late_f + maint_c

    # CREATE FINAL BILL OBJECT ***********************
    new_bill = Billing(bill_ref=bref,

                       tenant_name=tname,
                       room_no=rno,
                       room_cost=room_cost,
                       room_acc_cost=room_acc_cost,
                       electricity_cost=elec_cost,
                       water_cost=water_cost,
                       common_ser_cost=com_ser_cost,
                       other_ser_cost=oth_ser_cost,
                       overdue_amount=ovd_amt,
                       # -------------------
                       late_fee=late_f,
                       maint_cost=maint_c,
                       # -------------------

                       adjust=adj,
                       bill_total=total,

                       )

    # SAVE BILL OBJECT TO DB
    new_bill.save()


# END CREATE BILLS ====================================================================================

# START OF BILL PAYMENT ==============================================================================
@login_required
def month_bills(request):
    avail_open_bills = Billing.objects.filter(status='open').order_by('id')  # QS of OPEN Bills

    no_of_avail_open_bills = len(list(avail_open_bills))

    if no_of_avail_open_bills:
        latest_bill = Billing.objects.all().order_by('-bill_date').first()  # ASSUMING BILLS ARE AVAILABLE IN THE SYS.
        latest_bill_date = latest_bill.bill_date

        total_open_bill = len(Billing.objects.filter(bill_date__month=latest_bill_date.month))
        thai_bill_month = thaidate(latest_bill_date)['month']
        thai_bill_year = thaidate(latest_bill_date)['year']

    else:
        messages.info(request, 'ไม่มีบิลรอชำระ ...')
        return render(request, 'my_app/admin_page.html', {'section': 'month_bills'})

    return render(request, 'my_app/month_bills.html',
                  {
                      'section': 'month_bills',
                      'avail_open_bills': avail_open_bills,
                      'total_open_bill': total_open_bill,
                      'thai_bill_month': thai_bill_month,
                      'thai_bill_year': thai_bill_year,
                  })


@login_required
def pay_bill(request, bref):
    tenant_bill = get_object_or_404(Billing, bill_ref=bref, status='open')
    rmn = tenant_bill.room_no

    if request.method == 'POST':
        pay_form = PaymentForm(data=request.POST)

        if pay_form.is_valid():
            cd = pay_form.cleaned_data

            if (cd['payment_amount'] - tenant_bill.bill_total) > 0.5:  # compensate for round_up
                messages.error(request, 'ใส่ค่าไม่เหมาะสม ใส่ค่าใหม่ให้ถูกต้อง...')

                return redirect(reverse('pay_bill', kwargs={'bref': tenant_bill.bill_ref}))

            # --------------------------
            update_pf_and_bill(rmn, cd)
            # --------------------------

            messages.success(request, 'ห้อง {} ชำระค่าเช่าเรียบร้อยแล้ว'.format(rmn))
            return HttpResponseRedirect(reverse_lazy('month_bills'))

        else:
            messages.error(request, 'เกิดข้อผิดพลาดในการปิดบิล ห้อง {}'.format(tenant_bill.room_no))
            return HttpResponseRedirect(reverse_lazy('month_bills'))
    else:

        pay_form = PaymentForm()  # Blank form
        return render(request, 'my_app/pay_bill.html', {'section': 'month_bills', 'tenant_bill': tenant_bill, 'pay_form': pay_form})


# @login_required (cannot be used here !!!)
def update_pf_and_bill(roomno, cd):
    pf = get_object_or_404(TenantProfile, room_no__room_no=roomno)
    bill = get_object_or_404(Billing, room_no=roomno, status='open')

    # ----------- To ensure there will be no residual negative small value in OVD -------
    cf = bill.bill_total - cd['payment_amount']  # 3100.8 - 3101 = -.2
    if abs(cf) < 0.5:  # 3100.5-3101,3100.4-3100
        cf = 0
    bill.cf_amount = cf
    pf.cum_ovd = cf  # to combine with next month charge
    # ----------------------------------------------------------------------------------

    bill.payment_date = cd['payment_date']
    bill.payment_amount = cd['payment_amount']
    bill.status = 'close'

    # CALCULATE LATE-FEE COST TO UPDATE PF.LATE_FEE
    bill_month = bill.bill_date.month  # int
    pay_month = bill.payment_date.month

    if bill_month != pay_month:
        late_day = bill.payment_date.day - GV.LATE_DAY_MAX  # int
        if late_day < 0:
            late_day = 0
    else:
        late_day = 0

    late_fee = late_day * GV.LATE_FEE_PER_DAY

    pf.late_fee = late_fee  # to combine with next month charge

    # Update DB
    bill.save()
    pf.save()


# END OF BILL PAYMENT ================================================================================

@login_required
def report_type(request):
    return render(request, 'my_app/report_type.html', {'section': 'report'})


@login_required
def report_parameters(request):
    return render(request, 'my_app/report_parameters.html', {'section': 'report'})


@login_required
def monthly_report(request):
    closed_bills = Billing.objects.filter(status='close').order_by('-bill_date')
    if len(closed_bills) == 0:  # no closed bills (or no bills) in the system
        messages.error(request, 'ไม่มีบิลที่ชำระแล้วในระบบ !!!')
        return render(request, 'my_app/monthly_report.html', {'section': 'report', 'closed_bills': closed_bills})
    # ----------------------------------------------------------------------------------------
    report_m = int(request.POST['month'])
    report_y = int(request.POST['year'])
    report_date = date(report_y, report_m, 1)  # date obj
    thai_bd_m = thaidate(report_date)["month"]
    thai_bd_y = thaidate(report_date)["year"]

    open_bill_in_report_month = None
    total_bill_in_report_month = None
    closed_bill_in_report_month = None

    bld = request.POST['bld']
    if bld == 'AB':
        bld = 'A&B'

    if bld == 'A':
        open_bill_in_report_month = Billing.objects.filter(status='open', room_no__startswith='A', bill_date__month=report_date.month, bill_date__year=report_date.year)
        total_bill_in_report_month = Billing.objects.filter(bill_date__month=report_date.month, bill_date__year=report_date.year, room_no__startswith='A')
        closed_bill_in_report_month = Billing.objects.filter(status='close', room_no__startswith='A', bill_date__month=report_date.month, bill_date__year=report_date.year).order_by('room_no')

    elif bld == 'B':
        open_bill_in_report_month = Billing.objects.filter(status='open', room_no__startswith='B', bill_date__month=report_date.month, bill_date__year=report_date.year)
        total_bill_in_report_month = Billing.objects.filter(bill_date__month=report_date.month, bill_date__year=report_date.year, room_no__startswith='B')
        closed_bill_in_report_month = Billing.objects.filter(status='close', room_no__startswith='B', bill_date__month=report_date.month, bill_date__year=report_date.year).order_by('room_no')

    elif bld == 'A&B':
        open_bill_in_report_month = Billing.objects.filter(status='open', bill_date__month=report_date.month, bill_date__year=report_date.year).order_by('id')
        total_bill_in_report_month = Billing.objects.filter(bill_date__month=report_date.month, bill_date__year=report_date.year)
        closed_bill_in_report_month = Billing.objects.filter(status='close', bill_date__month=report_date.month, bill_date__year=report_date.year).order_by('room_no')

    trcac = 0
    tec = 0
    twc = 0
    tcsc = 0
    tosc = 0
    tovd = 0
    tlf = 0
    tmc = 0
    tbt = 0
    tpa = 0

    total_amt_list = []

    for bill in closed_bill_in_report_month:
        trcac += (bill.room_cost + bill.room_acc_cost + bill.adjust)
        tec += bill.electricity_cost
        twc += bill.water_cost
        tcsc += bill.common_ser_cost
        tosc += bill.other_ser_cost
        tovd += bill.overdue_amount
        tlf += bill.late_fee
        tmc += bill.maint_cost
        tbt += bill.bill_total
        tpa += bill.payment_amount
        total_amt_list = [trcac, tec, twc, tcsc, tosc, tovd, tlf, tmc, tbt, tpa]

    return render(request, 'my_app/monthly_report.html',
                  {
                      'section': "report",
                      'bld': bld,
                      'closed_bill_in_report_month': closed_bill_in_report_month,
                      'trcac': trcac,
                      'tec': tec,
                      'twc': twc,
                      'tcsc': tcsc,
                      'tosc': tosc,
                      'tovd': tovd,
                      'tlf': tlf,
                      'tmc': tmc,
                      'tbt': tbt,
                      'tpa': tpa,
                      'total_amt_list': total_amt_list,
                      'thai_bd_m': thai_bd_m,
                      'thai_bd_y': thai_bd_y,
                      'open_bill_in_report_month': open_bill_in_report_month,
                      'total_bill_in_report_month': total_bill_in_report_month,
                  })


@login_required
def extra_rates(request):
    extra = Extra.objects.all().order_by('-id')

    current_dt = datetime.now()

    return render(request, 'my_app/extra_rates.html', {'section': 'report', 'extra': extra, 'current_dt': current_dt})


@login_required
def elec_cpu_change(request):
    if request.method == 'POST':
        elec_cpu_form = Elec_cpu_change(request.POST)
        if elec_cpu_form.is_valid():
            cd = elec_cpu_form.cleaned_data

            ex_item = get_object_or_404(Extra, description='Electricity CPU')
            ex_item.cpu = cd['elec_cpu']
            ex_item.save()

            messages.info(request, 'ค่าไฟฟ้า ได้ถูกเปลี่ยนเป็นค่าใหม่เรียบร้อยแล้ว')

            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.error(request, 'มีข้อผิดพลาดเกิดขึ้น !!!')
    else:
        elec_cpu_form = Elec_cpu_change()
    return render(request, 'my_app/elec_cpu_change.html',
                  {
                      'section': 'report',
                      'elec_cpu_form': elec_cpu_form
                  })


@login_required
def water_cpu_change(request):
    if request.method == 'POST':
        water_cpu_form = Water_cpu_change(request.POST)
        if water_cpu_form.is_valid():
            cd = water_cpu_form.cleaned_data

            ex_item = get_object_or_404(Extra, description='Water CPU')
            ex_item.cpu = cd['water_cpu']
            ex_item.save()

            messages.info(request, 'ค่าน้ำ ได้ถูกเปลี่ยนเป็นค่าใหม่เรียบร้อยแล้ว')

            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.error(request, 'มีข้อผิดพลาดเกิดขึ้น !!!')
    else:
        water_cpu_form = Water_cpu_change()
    return render(request, 'my_app/water_cpu_change.html',
                  {
                      'section': 'report',
                      'water_cpu_form': water_cpu_form
                  })


@login_required
def room_type_rate(request):
    rm_type_rate = Room_type.objects.all()
    current_dt = datetime.now()

    return render(request, 'my_app/room_type_rate.html', {'section': 'report', 'rm_type_rate': rm_type_rate, 'current_dt': current_dt})


@login_required
def vacant_rooms(request):
    current_dt = datetime.now()

    all_room = Room.objects.all()  # QS
    cur_tn = TenantProfile.objects.all()  # QS
    oc_rm_set = []
    vac_rm_set = []
    for tn in cur_tn:  # tn is obj TenantProfile
        oc_rm_set.append(tn.room_no)  # "oc_rm_set" is a set of occupied Room obj
    for rm in all_room:  # rm is obj Room
        if rm not in oc_rm_set:
            vac_rm_set.append(rm)  # "vac_rm_set" is a set of vacant Room obj

    # Sort "vac_rm_set" ------------------
    rmn_set = []
    svac_rm_set = []
    for r in vac_rm_set:
        rmn_set.append(r.room_no)

    rmn_set.sort()
    # rmn_set.reverse()
    for r in rmn_set:
        for rmobj in vac_rm_set:
            if rmobj.room_no == r:
                svac_rm_set.append(rmobj)
    # -------------------------------------

    no_of_vac_room = len(vac_rm_set)

    return render(request, 'my_app/vacant_rooms.html',
                  {
                      'section': 'report',
                      'res': 'yes',
                      'svac_rm_set': svac_rm_set,
                      'current_dt': current_dt,
                      'no_of_vac_room': no_of_vac_room,

                  })


@login_required
def update_room_status(request, rmn):
    vac_rm = get_object_or_404(Room, room_no=rmn)  # room obj
    room_rate = vac_rm.room_type.rate

    if request.method == 'POST':
        book_room_form = BookRoomForm(data=request.POST)

        if book_room_form.is_valid():

            cd = book_room_form.cleaned_data

            vac_rm.status = cd['status']
            vac_rm.exmovein_date = cd['exmovein_date']
            vac_rm.save()

            messages.success(request, 'เปรี่ยนสถานะห้องว่าง {} เรียบร้อยแล้ว'.format(rmn))
            return render(request, 'my_app/report_type.html', {'section': 'report'})
        else:
            messages.error(request, 'เกิดข้อผิดพลาด!!! {}')
    else:
        book_room_form = BookRoomForm()  # Blank form
        return render(request, 'my_app/update_room_status.html', {'section': 'report', 'book_room_form': book_room_form, 'room_rate': room_rate, 'rmn': rmn})


@login_required
def current_tenants(request):
    cur_tenant_pfs = TenantProfile.objects.all().order_by('room_no')  # CHANGED 21 Jan 23

    total_tn = cur_tenant_pfs.count()

    current_dt = datetime.now()

    return render(request, 'my_app/current_tenants.html',
                  {
                      'section': 'report',

                      'cur_tenant_pfs': cur_tenant_pfs,
                      'current_dt': current_dt,
                      'total_tn': total_tn
                  })


@login_required
def misc_contents(request):
    return render(request, 'my_app/misc_contents.html', {'section': 'misc'})


# START OF CREATE NEW ACCOUNT USING EXCEL DATA *********************************************

@login_required
def pre_create_new_account_excel(request):
    # username_set, new_user_dict = create_new_account_subroutine()
    # no_of_new_user = len(username_set)

    new_tenant_d = create_new_account_subroutine()

    return render(request, 'my_app/pre_create_new_account_excel.html',
                  {
                      'section': 'misc',
                      'new_tenant_d': new_tenant_d,
                      'no_of_new_tenant': len(new_tenant_d.keys()),
                  })


@login_required
def create_new_account_excel(request):
    # EXISTING TENANTS *********************************************************************

    new_tenant_dict = create_new_account_subroutine()

    ref_extra = {9: "Electricity CPU", 10: 'Water CPU', 11: 'Garbage', 12: 'Bed&Mattress', 13: 'Clothing Cupboard', 14: 'Dressing Table', 15: 'TV Table', 16: 'Fridge', 17: 'Wifi', 18: 'Air-Conditioner', 19: 'Parking'}

    # Create a new tenant object **********************************************************
    no_of_new_tenants = len(new_tenant_dict.keys())

    if no_of_new_tenants > 0:
        number_of_new_tenant = 0
        for val in new_tenant_dict.values():
            # new_tenant_data = new_tenant_dict[i]
            username = val[1]
            password = val[2]
            first_name, last_name = val[3].split()

            new_tenant = CUser(username=username, first_name=first_name, last_name=last_name)

            # Set the chosen password
            new_tenant.set_password(password)

            # Save the new_tenant object
            new_tenant.save()

            # Create a new TenantProfile object *******************************************
            extra_set = []
            for j in range(9, 20):  # to cover extra items
                if val[j] == 1:
                    extra_set.append(ref_extra[j])

            tenant = new_tenant
            pin = val[4]
            phone = val[5]
            room_no = get_object_or_404(Room, room_no=val[0])
            start_date = val[8]
            end_date = start_date.__add__(timedelta(days=1 * 365))
            depo = val[6]
            adj = val[7]
            bill_ref = ''
            elec_unit = 0
            water_unit = 0

            new_tenanat_profile = TenantProfile(tenant=tenant, pin=pin, phone=phone, room_no=room_no, start_date=start_date, end_date=end_date, deposit=depo, adjust=adj, bill_ref=bill_ref, elec_unit=elec_unit, water_unit=water_unit)

            new_tenanat_profile.save()  # SAVE FIRST BEFORE ADDING EXTRA ITEMS

            for n in extra_set:
                new_tenanat_profile.extra.add(get_object_or_404(Extra, description=n))  # ADD ONE TO MANY ITEMS
            new_tenanat_profile.save()
            number_of_new_tenant += 1

        messages.info(request, f'ผู่้เช่าใหม่ {number_of_new_tenant} คน ถูกเอาเข้าระบบแล้ว')
    else:
        messages.info(request, f'ไม่มี ผู่้เช่าใหม่ ...')

    return render(request, 'my_app/create_new_account_excel.html', {'section': 'misc'})


def create_new_account_subroutine():
    # EXISTING TENANTS ************************************************************
    existing_tenant = CUser.objects.all().order_by('id')
    username_of_existing_tenants_set = []
    existing_tenant_dict = {}
    for i in existing_tenant:
        username_of_existing_tenants_set.append(i.username)
        existing_tenant_dict.update({i.username: i.first_name + " " + i.last_name})

    occupied_romno_set = []
    existing_tpf = TenantProfile.objects.all().order_by('room_no__room_no')
    for tpf in existing_tpf:
        occupied_romno_set.append(tpf.room_no.room_no)  # ['A302', 'B304',..]

    # Excel data range: min_row: 5, max_row: 56, min_col: 3, max_col: 22
    rmno_value_dict = get_excel_data(5, 56, 3, 22, GV.excel_file_new_ac)  # {rmno: [room_no,v1,..]}

    new_tenat_dict = {}
    for v in rmno_value_dict.values():
        rmno = v[0]
        if rmno is not None and rmno not in occupied_romno_set:
            new_tenat_dict.update({rmno: v})

    return new_tenat_dict


# END OF CREATE NEW ACCOUNT USING EXCEL DATA ******************************************


# START OF ISSUEING BILLS USING EXCEL *************************************************
@login_required
def issue_bill_excel(request):
    # Check fo OPEN Bills & abnormal conditios ----------------------------------------
    open_bill = Billing.objects.filter(status='open')  # QuerySet

    if open_bill:
        messages.error(request, 'มี {} บิล ยังค้างอยู่ ปิดให้เรียบร้อยก่อน'.format(len(open_bill)))
        return render(request, 'my_app/admin_page.html', {'section': 'misc'})
    else:
        all_bills = Billing.objects.all().order_by('-bill_date')  # QS
        if len(all_bills) != 0:
            tpfs = TenantProfile.objects.all()
            latest_bill = all_bills.first()  # bill obj
            latest_billdate = latest_bill.bill_date

            for i in tpfs:
                if i.bill_ref == "" and i.start_date < latest_billdate:
                    messages.error(request, 'ห้อง {} ไม่มี bill_ref !!! ... ตรวจสอบให้ถูกต้องก่อนดำเนินการต่อ ....'.format(i.room_no.room_no))
                    return render(request, 'my_app/admin_page.html', {'section': 'misc'})

    # use data from excel, data from tpf to create bills -----------------------------------
    bld1_elec_water_consump_dict = get_excel_data(5, 23, 2, 8, GV.excel_file_ew_con)
    bld2_elec_water_consump_dict = get_excel_data(7, 22, 10, 16, GV.excel_file_ew_con)

    # Create Bills for each tetant ----------------------------------------------------------
    existing_tpfs = TenantProfile.objects.all().order_by('room_no__room_no')  # QS
    no_of_bill = 0
    for tpf in existing_tpfs:  # tpf by tpf

        tn_rmno = tpf.room_no.room_no  # 'A105' etc.  str

        elec_water_concumption_dict = None

        if tn_rmno[0] == 'A':  # SLICING "A106"
            elec_water_concumption_dict = bld1_elec_water_consump_dict
        elif tn_rmno[0] == 'B':
            elec_water_concumption_dict = bld2_elec_water_consump_dict

        tpf_multi_factor = 1

        if tpf.bill_ref == "":
            if tpf.start_date.month == date.today().month:
                tpf_multi_factor = decimal.Decimal((calendar.monthrange(date.today().year, date.today().month)[1] - tpf.start_date.day) / 30)
            else:
                tpf_multi_factor = decimal.Decimal((calendar.monthrange(date.today().year, date.today().month)[1] + (calendar.monthrange(tpf.start_date.year, tpf.start_date.month)[1] - tpf.start_date.day)) / 30)

        bill_ref_str = get_ref_string()
        tpf.bill_ref = bill_ref_str

        # CREATE NEW BILL OBJ. for this tenant *************************************************
        new_bill = Billing(bill_ref=bill_ref_str, tenant_name=tpf.tenant, room_no=tpf.room_no)

        # new_bill.room_cost = tpf.room_no.room_type.rate * tpf_multi_factor
        new_bill.room_cost = tpf.room_no.room_type.rate * tpf_multi_factor

        tpf_extras = tpf.extra.all()  # QS

        rm_ac_cost = 0
        osc = 0
        elec_cpu = 0
        water_cpu = 0

        for e in tpf_extras:  # one extra obj.
            if e.description in ['Bed&Mattress', 'Clothing Cupboard', 'Dressing Table', 'TV Table', 'Fridge', 'Air-Conditioner']:
                rm_ac_cost += e.cpu
            elif e.description in ['Garbage', 'Wifi', 'Parking']:
                osc += e.cpu
            elif e.description == 'Electricity CPU':
                elec_cpu = e.cpu
            elif e.description == 'Water CPU':
                water_cpu = e.cpu

        new_bill.room_acc_cost = rm_ac_cost * tpf_multi_factor

        tpf.elec_unit = elec_water_concumption_dict[tn_rmno][3]
        tpf.water_unit = elec_water_concumption_dict[tn_rmno][6]

        new_bill.electricity_cost = elec_cpu * tpf.elec_unit
        new_bill.water_cost = water_cpu * tpf.water_unit

        new_bill.common_ser_cost = elec_water_concumption_dict[tn_rmno][3] * GV.COMMOM_SERVICE_CPU  # No Charge

        new_bill.other_ser_cost = osc

        new_bill.overdue_amount = tpf.cum_ovd
        new_bill.adjust = tpf.adjust

        new_bill.late_fee = tpf.late_fee
        new_bill.maint_cost = tpf.maint_cost

        # RESET late_fee and maint_cost in tpf
        tpf.late_fee = 0
        tpf.maint_cost = 0

        bill_total = new_bill.room_cost + new_bill.room_acc_cost + new_bill.electricity_cost + new_bill.water_cost + new_bill.common_ser_cost + new_bill.other_ser_cost + new_bill.overdue_amount + new_bill.adjust + new_bill.late_fee + new_bill.maint_cost

        new_bill.bill_total = bill_total

        # SAVE tpf and new bill for each tenant
        tpf.save()
        new_bill.save()

        no_of_bill += 1

    messages.info(request, 'Total {} bills issued.'.format(no_of_bill))

    return render(request, 'my_app/issue_bill_excel.html',
                  {
                      'section': 'misc',
                      'bld1_elec_water_consump_dict': bld1_elec_water_consump_dict,
                      'bld2_elec_water_consump_dict': bld2_elec_water_consump_dict,
                  })


# END OF ISSUEING BILLS USING EXCEL **************************************************************

# START OF DELETE USERS ==============================================================================
@login_required
def manage_users(request):
    return render(request, 'my_app/manage_users.html', {'section': 'misc'})


@login_required
def user_list_to_delete(request):
    query_set_tenantprofile, sorted_normal_tenantprofile_dict = list_existing_users(request)

    current_date_time = datetime.now()

    return render(request, 'my_app/user_list_to_delete.html',
                  {
                      'section': 'misc',
                      'tenantprofiles': query_set_tenantprofile,
                      'dict': sorted_normal_tenantprofile_dict,
                      'current_date_time': current_date_time
                  })


@login_required
def confirm_delete_user(request, k):
    tprofile = TenantProfile.objects.get(room_no__room_no=k)

    rmn = tprofile.room_no.room_no
    name = tprofile.tenant.first_name + " " + tprofile.tenant.last_name

    return render(request, 'my_app/confirm_delete_users.html', {'section': 'misc', 'rmn': rmn, 'name': name})


@login_required
def delete_user(request, rmn):
    tprofile = TenantProfile.objects.get(room_no__room_no=rmn)
    user = tprofile.tenant

    b_ref = tprofile.bill_ref

    try:

        tbill = get_object_or_404(Billing, bill_ref=b_ref)

        b_status = tbill.status

        if b_status == 'close':
            user.delete()
            messages.success(request, '{} ห้อง {} ถูกลบออกจากระบบแล้ว !!!'.format(user, rmn))

        else:
            messages.error(request, 'บิลยังค้างอยู่ ไม่สามารถลบได้ ... ต้องปิดบิลก่อน !!!')

    except:
        user.delete()
        messages.success(request, '{} ห้อง {} ถูกลบออกจากระบบแล้ว !!!'.format(user, rmn))

    current_date_time = datetime.now()

    query_set_tenantprofile, sorted_normal_tenantprofile_dict = list_existing_users(request)

    return render(request, 'my_app/user_list_to_delete.html',
                  {
                      'section': 'misc',

                      'tenantprofiles': query_set_tenantprofile,
                      'dict': sorted_normal_tenantprofile_dict,
                      'current_date_time': current_date_time,
                  })


def list_existing_users(request):
    query_set_tenantprofile = TenantProfile.objects.all().order_by('room_no')

    tenantprofile_dict = {}

    for i in query_set_tenantprofile:
        name = i.tenant.first_name + " " + i.tenant.last_name
        rmn = i.room_no.room_no
        phone = i.phone
        name_phone = name + " " + '(' + phone + ')'
        tenantprofile_dict.update({rmn: name_phone})  # {'A105': 'Ratchada R.', ....}

    sorted_normal_tenantprofile_dict = tenantprofile_dict

    return query_set_tenantprofile, sorted_normal_tenantprofile_dict


class Register(generic.CreateView):
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('register_done')
    template_name = 'registration/register.html'


def register_done(request):
    return render(request, 'registration/register_done.html')


@login_required
def change_password(request):
    return render(request, 'my_app/change_password.html')


# END OF DELETE USERS ================================================================================

def maintenance_charge(request):
    occupied_rm = []
    tpfs = TenantProfile.objects.all()
    for tp in tpfs:
        occupied_rm.append(tp.room_no.room_no)
    occupied_rm.sort()

    if request.method == 'POST':
        rn = request.POST.get('room_no', '')
        jr = get_ref_string()

        jc = 0
        try:
            jc = decimal.Decimal((request.POST.get('job_cost', 0)))
        except:

            messages.error(request, 'ใส่ค่าไม่ถูกต้อง !!!')
            return render(request, 'my_app/maintenanace_charge.html', {'section': 'misc', 'occupied_rm': occupied_rm})

        pf = get_object_or_404(TenantProfile, room_no__room_no=rn)
        pf.maint_cost += jc

        pf.save()

        ms = MaintenanceService(job_ref=jr, room_no=rn, job_cost=jc)
        ms.save()
        messages.success(request, 'ค่าบริการ ได้ถูกคิดรวมเป็นค่าใช้จ่าย ห้อง {} เรียบร้อยแล้ว'.format(rn))

        return render(request, 'my_app/admin_page.html', {'section': 'misc'})

    return render(request, 'my_app/maintenanace_charge.html',
                  {
                      'section': 'misc',
                      'occupied_rm': occupied_rm,
                  })


# START OF TENANT PART  ==============================================================================

@login_required
def new_tenant(request):
    tenant_name = str(request.user)
    fn, ln = tenant_name.split(" ")

    section = 'tenant_bill'

    return render(request, 'my_app/new_tenant.html', {'fn': fn, 'section': section})


@login_required
def tenant_profile(request):
    usr = str(request.user)
    fn, ln = usr.split(" ")

    try:
        tenant_pf = TenantProfile.objects.get(tenant__first_name=fn, tenant__last_name=ln)
    except Exception as err:
        messages.error(request, 'ERROR: {} '.format(str(err)))
        messages.error(request, 'มีข้อผิดพลาดเกิดขึ้น !!!')

        return HttpResponseRedirect(reverse_lazy('login'))
    else:
        exd = {}
        exd.setdefault('Electricity CPU', 0)
        exd.setdefault('Water CPU', 0)
        exd.setdefault('Garbage', 0)
        exd.setdefault('Parking', 0)
        exd.setdefault('Wifi', 0)

        exd.setdefault('Bed&Mattress', 0)

        exd.setdefault('Dressing Table', 0)
        exd.setdefault('Clothing Cupboard', 0)
        exd.setdefault('TV Table', 0)
        exd.setdefault('Fridge', 0)
        exd.setdefault('Air-Conditioner', 0)

        for e in tenant_pf.extra.all():
            exd.update({e.description: e.cpu})

        room_acc_cost = exd['Bed&Mattress'] + exd['Dressing Table'] \
                        + exd['Clothing Cupboard'] + exd['TV Table'] + exd['Fridge'] \
                        + exd['Air-Conditioner']

        oth_ser_cost = exd['Garbage'] + exd['Parking'] + exd['Wifi']

        cur_dt = datetime.now()

        return render(request, 'my_app/tenant_profile.html',
                      {'section': 'tenant_profile', 'tenant_pf': tenant_pf, 'room_acc_cost': room_acc_cost,
                       'oth_ser_cost': oth_ser_cost, 'cur_dt': cur_dt})


@login_required
def tenant_bill(request):
    t_name = str(request.user)
    fn, ln = t_name.split(' ')

    tpf = get_object_or_404(TenantProfile, tenant__first_name=fn, tenant__last_name=ln)

    active_bill_ref = tpf.bill_ref

    try:
        tn_bill = get_object_or_404(Billing, bill_ref=active_bill_ref)
    except:
        # *** NEW TENANT ***
        return HttpResponseRedirect(reverse_lazy('new_tenant'))

    room_with_acc_cost = tn_bill.room_cost + tn_bill.room_acc_cost + tn_bill.adjust
    bill_misc = tn_bill.late_fee + tn_bill.maint_cost
    bill_total = tn_bill.bill_total

    pay_date = tn_bill.payment_date
    pay_amt = tn_bill.payment_amount

    bill_m = tn_bill.bill_date.month
    bill_m_th = get_thai_month_name_int(bill_m)

    bill_y = tn_bill.bill_date.year
    bill_y_th = get_thai_year_int(bill_y)

    next_m = bill_m + 1
    if next_m > 12:
        next_m = 1
    next_m_th = get_thai_month_name_int(next_m)  # January, February, ...

    next_y = bill_y
    if bill_m == 12:
        next_y = bill_y + 1
    next_y_th = get_thai_year_int(next_y)

    if tn_bill.status == 'open':
        paid_str = 'รอชำระ'
    else:
        paid_str = 'ชำระแล้ว ณ วันที่ {0} {1} {2} จำนวน {3:,.0f} บาท'.format(pay_date.day, get_thai_month_name_str(str(pay_date)), get_thai_year_str(str(pay_date)), pay_amt)

    return render(request, 'my_app/tenant_bill.html',
                  {
                      'section': 'tenant_bill',
                      'tn_bill': tn_bill,

                      'room_with_acc_cost': room_with_acc_cost,
                      'bill_misc': bill_misc,
                      'bill_total': bill_total,

                      'bill_m_th': bill_m_th,
                      'bill_y_th': bill_y_th,

                      'next_m_th': next_m_th,
                      'next_y_th': next_y_th,

                      'paid_str': paid_str,

                  })


def tenant_info(request):
    return render(request, 'my_app/tenant_info.html', {'section': 'tenant_info'})


# END OF TENANT PART =============================================================================

# START OF MISC. *********************************************************************************
def get_excel_data(ws_min_r, ws_max_r, ws_min_c, ws_max_c, excel_f):
    os.chdir(GV.excel_file_path)
    wb = load_workbook(excel_f, data_only=True)  # TO WORK WITH VALUES ONLY
    ws = wb.active

    ws_min_row = ws_min_r
    ws_max_row = ws_max_r

    ws_min_col = ws_min_c
    ws_max_col = ws_max_c

    rmno_value_dict = {}
    for row in range(ws_min_row, ws_max_row):
        ws_range = ws.iter_cols(min_row=row, max_row=row, min_col=ws_min_col, max_col=ws_max_col, values_only=True)  # one row at a time
        temporary_set = []
        for cell in ws_range:  # ws_range obj: <ws_range: room_no,...,....>
            for e in cell:
                temporary_set.append(e)  # [room_no,?,?, ....]
        rmno_value_dict.update({temporary_set[0]: temporary_set})  # {'A105': [rmno,username, passwsord,..]}

    return rmno_value_dict  # {'A105': [rmno,username, passwsord,..]}


def get_ref_string():
    char_str = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    random.shuffle(char_str)
    fd = random.choice(char_str)

    sd = str(random.randint(0, 9)) + str(random.randint(0, 9)) + str(random.randint(0, 9)) + str(
        random.randint(0, 9))
    ref_str = fd + '-' + sd

    return ref_str


def get_eng_month_name(m: int):
    md = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
          9: 'September',
          10: 'October', 11: 'November', 12: 'December'}
    im = int(m)
    return md[im]


def get_thai_month_name_str(bill_date: str):
    md = {1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฏาคม',
          8: 'สิงหาคม', 9: 'กันยายน',
          10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'}

    y, m, d = bill_date.split('-')

    im = int(m)
    return md[im]


def get_thai_month_name_int(month: int):
    md = {1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฏาคม',
          8: 'สิงหาคม', 9: 'กันยายน',
          10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'}

    return md[month]


def get_thai_year_str(bill_date: str):
    y, m, d = bill_date.split('-')

    christ_y = int(y)
    buddist_y = christ_y + 543

    return str(buddist_y)


def get_thai_year_int(year: int):
    buddist_y = year + 543

    return str(buddist_y)


def make_date_string(self, ds: str):
    y, m, d = str(ds).split('-')
    return d + '-' + m + '-' + y


def give_error_message(error_msg):
    print(error_msg)


def give_info_message(error_msg):
    print(error_msg)


def get_aware_datetime(date_str):
    ret = parse_datetime(date_str)
    if not is_aware(ret):
        ret = make_aware(ret)
    return ret


from django import template

register = template.Library


def get_thai_month_buddist_year(bill_date: str):
    md = {1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฏาคม',
          8: 'สิงหาคม', 9: 'กันยายน',
          10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'}

    y, m, d = bill_date.split('-')

    thai_month = md[int(m)]

    christ_y = int(y)
    buddist_y = christ_y + 543

    return thai_month, buddist_y


def thaidate(date_obj):
    date_dict = {}
    n = ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฏาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
    d = date_obj.day
    m = n[date_obj.month - 1]
    y = date_obj.year + 543

    date_dict.update({'day': d})
    date_dict.update({'month': m})
    date_dict.update({'year': y})

    return date_dict


def get_thai_year(bill_date: str):
    y, m, d = bill_date.split('-')

    christ_y = int(y)
    buddist_y = christ_y + 543

    return str(buddist_y)

# END OF MISC. *****************************************************************************
